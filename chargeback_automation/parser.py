from __future__ import annotations

import csv
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .chargeback_types import FIELD_TO_HEADER


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", normalize_text(value))


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


ALIASES: dict[str, list[str]] = {
    "nsu": ["nsu", "numero nsu", "cod nsu", "codigo nsu"],
    "data_transacao": [
        "data da transacao",
        "data transacao",
        "dt transacao",
        "data da compra",
        "data compra",
    ],
    "data_abertura_chargeback": [
        "data abertura de chargeback",
        "data de abertura de chargeback",
        "abertura chargeback",
        "data abertura contestacao",
        "data de abertura da contestacao",
    ],
    "prazo_contestacao": [
        "prazo de contestacao",
        "prazo contestacao",
        "data limite",
        "limite contestacao",
        "vencimento contestacao",
    ],
    "valor_chargeback": [
        "valor do chargeback",
        "valor chargeback",
        "valor contestado",
        "valor da contestacao",
        "valor",
    ],
    "bandeira": ["bandeira", "cartao bandeira", "bandeira cartao"],
    "valor_taxa": ["valor taxa", "valor da taxa", "taxa", "fee"],
    "transportadora": ["transportadora", "logistica", "carrier"],
    "numero_rastreio": [
        "numero do rastreio",
        "número do rastreio",
        "rastreio",
        "codigo rastreio",
        "cod rastreio",
        "tracking",
        "tracking number",
    ],
    "data_envio_cliente": [
        "data de envio para o cliente",
        "data envio cliente",
        "data envio",
        "envio cliente",
        "data postagem",
    ],
    "retorno_aprovacao": [
        "retorno de aprovacao",
        "retorno aprovação",
        "aprovacao",
        "retorno",
        "status aprovacao",
    ],
    "id_interno": ["id interno", "id", "pedido", "pedido interno", "id pedido"],
    "motivo_recusa": ["motivo de recusa", "motivo recusa", "recusa", "motivo"],
    "obs": ["obs", "observacao", "observacoes", "observação", "comentario"],
    "acao": ["acao", "ação", "proxima acao manual", "tratativa"],
    "aprovacao_por_quem": [
        "aprovacao por quem",
        "aprovação por quem?",
        "aprovacao por quem?",
        "responsavel",
        "responsável",
        "aprovador",
    ],
    "classificacao_ia": ["classificacao ia", "classificação ia", "classificacao"],
    "risco": ["risco"],
    "pendencias": ["pendencias", "pendências"],
    "proxima_acao": ["proxima acao", "próxima ação", "acao recomendada"],
    "status_prazo": ["status prazo", "status do prazo"],
    "status_vtex": ["status vtex", "status pedido", "status do pedido"],
    "data_faturamento": ["data faturamento", "faturamento", "data da nota", "data nf"],
    "data_analise": ["data analise", "data análise", "analisado em"],
    "origem_classificacao": [
        "origem classificacao",
        "origem classificação",
        "classification source",
    ],
}

NORMALIZED_ALIAS_TO_FIELD: dict[str, str] = {}
for field, header in FIELD_TO_HEADER.items():
    NORMALIZED_ALIAS_TO_FIELD[normalize_header(header)] = field
for field, aliases in ALIASES.items():
    for alias in aliases:
        NORMALIZED_ALIAS_TO_FIELD[normalize_header(alias)] = field


def field_for_header(header: Any) -> str | None:
    return NORMALIZED_ALIAS_TO_FIELD.get(normalize_header(header))


def parse_money(value: Any) -> float | Any:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip()
    cleaned = re.sub(r"[^\d,.\-]", "", text)
    if cleaned in {"", "-", ".", ","}:
        return value
    try:
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        return float(cleaned)
    except ValueError:
        return value


def parse_date(value: Any) -> date | Any:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return value

    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return value


def normalize_record(raw: dict[str, Any]) -> dict[str, Any]:
    record: dict[str, Any] = {}
    for header, value in raw.items():
        field = field_for_header(header)
        if field:
            record[field] = value

    for money_field in ("valor_chargeback", "valor_taxa"):
        if money_field in record:
            record[money_field] = parse_money(record[money_field])
    for date_field in (
        "data_transacao",
        "data_abertura_chargeback",
        "prazo_contestacao",
        "data_envio_cliente",
        "data_faturamento",
        "data_analise",
    ):
        if date_field in record:
            record[date_field] = parse_date(record[date_field])
    return record


def row_is_empty(values: Iterable[Any]) -> bool:
    return all(is_blank(value) for value in values)


def parse_csv(path: str | Path) -> list[dict[str, Any]]:
    file_path = Path(path)
    raw_bytes = file_path.read_bytes()
    for encoding in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            text = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = raw_bytes.decode("utf-8", errors="replace")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";"

    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return [normalize_record(row) for row in reader if not row_is_empty(row.values())]


def find_header_row_and_map(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    best_row_index = 0
    best_map: dict[str, int] = {}
    for index, row in enumerate(rows):
        current: dict[str, int] = {}
        for col_index, header in enumerate(row, start=1):
            field = field_for_header(header)
            if field and field not in current:
                current[field] = col_index
        if len(current) > len(best_map):
            best_row_index = index
            best_map = current
    return best_row_index + 1, best_map


def parse_xlsx(path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    probe_rows = list(
        worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row), values_only=True)
    )
    header_row, mapping = find_header_row_and_map(probe_rows)
    records: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if row_is_empty(row):
            continue
        raw = {}
        for field, col_index in mapping.items():
            raw[FIELD_TO_HEADER[field]] = row[col_index - 1] if col_index - 1 < len(row) else None
        records.append(normalize_record(raw))
    workbook.close()
    return records


def parse_file(path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return parse_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx(path, sheet_name=sheet_name)
    raise ValueError("Formato não suportado. Use CSV ou XLSX.")
