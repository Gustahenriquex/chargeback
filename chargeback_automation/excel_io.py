from __future__ import annotations

import csv
import shutil
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .chargeback_types import (
    CLASSIFICATIONS,
    DEADLINE_STATUSES,
    EDITABLE_FIELDS,
    FIELD_TO_HEADER,
    GENERATED_COLUMNS,
    NEXT_ACTIONS,
    ORIGINAL_COLUMNS,
    RISKS,
)
from .classifier import checklist_for_record, classify_record
from .parser import field_for_header, is_blank, normalize_record, parse_date, row_is_empty

SUMMARY_SHEET = "Resumo Chargeback"
CHECKLIST_SHEET = "Checklist por Caso"
LISTS_SHEET = "_listas_chargeback"

GREEN_FILL = "D9EAD3"
YELLOW_FILL = "FFF2CC"
RED_FILL = "F4CCCC"
BLUE_FILL = "D9EAF7"
HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
GRID = Side(style="thin", color="D9E2F3")


def open_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Base Chargeback"
    sheet.append(ORIGINAL_COLUMNS + GENERATED_COLUMNS)
    return workbook


def workbook_permission_error(path: Path) -> RuntimeError:
    return RuntimeError(
        "Não foi possível acessar a planilha. Feche o arquivo no Excel/OneDrive "
        f"e tente novamente: {path}"
    )


def find_header_row(ws) -> int:
    best_row = 1
    best_score = 0
    max_probe = min(ws.max_row or 1, 30)
    for row_index in range(1, max_probe + 1):
        score = 0
        for cell in ws[row_index]:
            if field_for_header(cell.value):
                score += 1
        if score > best_score:
            best_score = score
            best_row = row_index
    return best_row


def column_map(ws, header_row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[header_row]:
        field = field_for_header(cell.value)
        if field and field not in mapping:
            mapping[field] = cell.column
    return mapping


def choose_data_sheet(workbook: Workbook, requested: str | None = None):
    if requested:
        if requested not in workbook.sheetnames:
            return workbook.create_sheet(requested)
        return workbook[requested]

    best_sheet = workbook.active
    best_score = -1
    for ws in workbook.worksheets:
        if ws.title in {SUMMARY_SHEET, CHECKLIST_SHEET, LISTS_SHEET}:
            continue
        header_row = find_header_row(ws)
        score = len(column_map(ws, header_row))
        if score > best_score:
            best_score = score
            best_sheet = ws
    return best_sheet


def ensure_columns(ws, header_row: int) -> dict[str, int]:
    mapping = column_map(ws, header_row)
    wanted_headers = ORIGINAL_COLUMNS + GENERATED_COLUMNS
    for header in wanted_headers:
        field = field_for_header(header)
        if field and field not in mapping:
            next_col = ws.max_column + 1
            ws.cell(row=header_row, column=next_col, value=header)
            mapping[field] = next_col
    return mapping


def record_from_row(ws, row_index: int, mapping: dict[str, int]) -> dict[str, Any]:
    raw = {}
    for field, col_index in mapping.items():
        header = FIELD_TO_HEADER.get(field)
        if header:
            raw[header] = ws.cell(row=row_index, column=col_index).value
    return normalize_record(raw)


def append_records(ws, header_row: int, records: list[dict[str, Any]]) -> None:
    if not records:
        return
    mapping = ensure_columns(ws, header_row)
    start_row = max(ws.max_row + 1, header_row + 1)
    for offset, record in enumerate(records):
        row_index = start_row + offset
        for field, value in record.items():
            if field in mapping:
                ws.cell(row=row_index, column=mapping[field], value=value)


def empty_data_row(ws, row_index: int, mapping: dict[str, int]) -> bool:
    original_fields = [field_for_header(header) for header in ORIGINAL_COLUMNS]
    values = [
        ws.cell(row=row_index, column=mapping[field]).value
        for field in original_fields
        if field and field in mapping
    ]
    return row_is_empty(values)


def classify_sheet(ws, header_row: int, near_days: int, today: date) -> list[dict[str, Any]]:
    mapping = ensure_columns(ws, header_row)
    rows: list[dict[str, Any]] = []
    for row_index in range(header_row + 1, ws.max_row + 1):
        if empty_data_row(ws, row_index, mapping):
            continue
        record = record_from_row(ws, row_index, mapping)
        analysis = classify_record(record, today=today, near_days=near_days)
        for field, value in analysis.as_record().items():
            ws.cell(row=row_index, column=mapping[field], value=value)
        rows.append(
            {
                "row_index": row_index,
                "record": record,
                "analysis": analysis,
                "generated": analysis.as_record(),
            }
        )
    return rows


def safe_sheet_name(workbook: Workbook, name: str) -> str:
    if name not in workbook.sheetnames:
        return name
    return name


def remove_generated_sheet(workbook: Workbook, name: str) -> None:
    if name in workbook.sheetnames:
        workbook.remove(workbook[name])


def build_summary_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    remove_generated_sheet(workbook, SUMMARY_SHEET)
    ws = workbook.create_sheet(SUMMARY_SHEET, 0)
    ws.sheet_view.showGridLines = False

    total_cases = len(rows)
    status_counter = Counter(row["analysis"].status_prazo for row in rows)
    high_risk = sum(1 for row in rows if row["analysis"].risco == "Alto")
    no_tracking = sum(
        1 for row in rows if "Sem rastreio informado" in row["analysis"].pendencias
    )
    total_value = 0.0
    for row in rows:
        value = row["record"].get("valor_chargeback")
        if isinstance(value, (int, float)):
            total_value += float(value)

    cards = [
        ("Total de casos", total_cases),
        ("Dentro do prazo", status_counter["Dentro do prazo"]),
        ("Próximos do vencimento", status_counter["Próximo do vencimento"]),
        ("Vencidos", status_counter["Vencido"]),
        ("Valor total contestado", total_value),
        ("Casos sem rastreio", no_tracking),
        ("Risco alto", high_risk),
    ]

    ws["A1"] = "Checklist Chargeback & Contestação"
    ws["A1"].font = Font(bold=True, size=18, color="1F2937")
    ws.merge_cells("A1:G1")
    ws["A2"] = f"Atualizado em {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(color="64748B")

    for index, (label, value) in enumerate(cards):
        col = index + 1
        cell_label = ws.cell(row=4, column=col, value=label)
        cell_value = ws.cell(row=5, column=col, value=value)
        cell_label.font = Font(bold=True, color="FFFFFF")
        cell_label.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell_label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_value.font = Font(bold=True, size=14, color="111827")
        cell_value.fill = PatternFill("solid", fgColor="F8FAFC")
        cell_value.alignment = Alignment(horizontal="center", vertical="center")
        cell_label.border = Border(top=GRID, left=GRID, right=GRID, bottom=GRID)
        cell_value.border = Border(top=GRID, left=GRID, right=GRID, bottom=GRID)
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws["E5"].number_format = 'R$ #,##0.00'

    ws["A8"] = "Distribuição por classificação"
    ws["A8"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Classificação IA", "Quantidade"])
    classification_counter = Counter(row["analysis"].classificacao_ia for row in rows)
    for name, count in classification_counter.most_common():
        ws.append([name, count])

    ws["D8"] = "Distribuição por risco"
    ws["D8"].font = Font(bold=True, size=12)
    ws["D10"] = "Risco"
    ws["E10"] = "Quantidade"
    risk_counter = Counter(row["analysis"].risco for row in rows)
    start = 11
    for offset, risk in enumerate(RISKS):
        ws.cell(row=start + offset, column=4, value=risk)
        ws.cell(row=start + offset, column=5, value=risk_counter[risk])

    for row in ws.iter_rows(min_row=10, max_row=max(ws.max_row, 13), min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(top=GRID, left=GRID, right=GRID, bottom=GRID)
            if cell.row == 10 and cell.value:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=HEADER_FILL)


def build_checklist_sheet(workbook: Workbook, rows: list[dict[str, Any]], today: date) -> None:
    remove_generated_sheet(workbook, CHECKLIST_SHEET)
    ws = workbook.create_sheet(CHECKLIST_SHEET)
    headers = [
        "Linha base",
        "NSU",
        "ID interno",
        "Classificação IA",
        "Risco",
        "Status prazo",
        "Próxima ação",
        "Item checklist",
        "Status",
        "Detalhe",
        "Observação manual",
        "Responsável/aprovação",
    ]
    ws.append(headers)
    for row in rows:
        record = row["record"]
        analysis = row["analysis"]
        for item, status, detail in checklist_for_record(record, analysis, today=today):
            ws.append(
                [
                    row["row_index"],
                    record.get("nsu"),
                    record.get("id_interno"),
                    analysis.classificacao_ia,
                    analysis.risco,
                    analysis.status_prazo,
                    analysis.proxima_acao,
                    item,
                    status,
                    detail,
                    record.get("obs"),
                    record.get("aprovacao_por_quem"),
                ]
            )
    style_table(ws, header_row=1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column, width in {
        "A": 12,
        "B": 18,
        "C": 18,
        "D": 24,
        "E": 12,
        "F": 22,
        "G": 30,
        "H": 42,
        "I": 12,
        "J": 45,
        "K": 35,
        "L": 26,
    }.items():
        ws.column_dimensions[column].width = width


def build_lists_sheet(workbook: Workbook) -> None:
    remove_generated_sheet(workbook, LISTS_SHEET)
    ws = workbook.create_sheet(LISTS_SHEET)
    lists = {
        "A": ["Classificação IA", *CLASSIFICATIONS],
        "B": ["Risco", *RISKS],
        "C": ["Status prazo", *DEADLINE_STATUSES],
        "D": ["Próxima ação", *NEXT_ACTIONS],
        "E": ["Retorno de aprovação", "Aprovado", "Recusado", "Em análise", "Pendente"],
    }
    for col, values in lists.items():
        for row_index, value in enumerate(values, start=1):
            ws[f"{col}{row_index}"] = value
    ws.sheet_state = "hidden"


def add_validations(workbook: Workbook, ws, mapping: dict[str, int], header_row: int) -> None:
    if LISTS_SHEET not in workbook.sheetnames:
        build_lists_sheet(workbook)
    max_row = max(ws.max_row, header_row + 1)
    validations = {
        "classificacao_ia": f"={LISTS_SHEET}!$A$2:$A${len(CLASSIFICATIONS) + 1}",
        "risco": f"={LISTS_SHEET}!$B$2:$B${len(RISKS) + 1}",
        "status_prazo": f"={LISTS_SHEET}!$C$2:$C${len(DEADLINE_STATUSES) + 1}",
        "proxima_acao": f"={LISTS_SHEET}!$D$2:$D${len(NEXT_ACTIONS) + 1}",
        "retorno_aprovacao": f"={LISTS_SHEET}!$E$2:$E$5",
    }
    for field, formula in validations.items():
        if field not in mapping:
            continue
        letter = get_column_letter(mapping[field])
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}{header_row + 1}:{letter}{max_row}")


def style_table(ws, header_row: int) -> None:
    max_col = ws.max_column
    max_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=GRID, left=GRID, right=GRID, bottom=GRID)

    for row in ws.iter_rows(min_row=header_row + 1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = Border(top=GRID, left=GRID, right=GRID, bottom=GRID)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    ws.row_dimensions[header_row].height = 28
    for row_index in range(header_row + 1, max_row + 1):
        ws.row_dimensions[row_index].height = 20


def apply_conditional_formatting(ws, mapping: dict[str, int], header_row: int) -> None:
    max_row = max(ws.max_row, header_row + 1)
    fills = {
        "green": PatternFill("solid", fgColor=GREEN_FILL),
        "yellow": PatternFill("solid", fgColor=YELLOW_FILL),
        "red": PatternFill("solid", fgColor=RED_FILL),
        "blue": PatternFill("solid", fgColor=BLUE_FILL),
    }
    rules = {
        "risco": [
            ('="Baixo"', fills["green"]),
            ('="Médio"', fills["yellow"]),
            ('="Alto"', fills["red"]),
        ],
        "status_prazo": [
            ('="Dentro do prazo"', fills["green"]),
            ('="Próximo do vencimento"', fills["yellow"]),
            ('="Vencido"', fills["red"]),
            ('="Sem prazo informado"', fills["yellow"]),
        ],
        "retorno_aprovacao": [
            ('="Aprovado"', fills["green"]),
            ('="Recusado"', fills["red"]),
            ('="Pendente"', fills["yellow"]),
        ],
    }
    for field, field_rules in rules.items():
        if field not in mapping:
            continue
        letter = get_column_letter(mapping[field])
        target = f"{letter}{header_row + 1}:{letter}{max_row}"
        for formula_suffix, fill in field_rules:
            formula = f"{letter}{header_row + 1}{formula_suffix}"
            ws.conditional_formatting.add(target, FormulaRule(formula=[formula], fill=fill))

    if "pendencias" in mapping:
        letter = get_column_letter(mapping["pendencias"])
        target = f"{letter}{header_row + 1}:{letter}{max_row}"
        ws.conditional_formatting.add(
            target,
            FormulaRule(formula=[f'LEN({letter}{header_row + 1})>0'], fill=fills["red"]),
        )


def apply_layout(ws, mapping: dict[str, int], header_row: int, workbook: Workbook) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions
    style_table(ws, header_row=header_row)
    apply_conditional_formatting(ws, mapping, header_row=header_row)
    add_validations(workbook, ws, mapping, header_row=header_row)
    widths = {
        "nsu": 18,
        "data_transacao": 16,
        "data_abertura_chargeback": 24,
        "prazo_contestacao": 20,
        "valor_chargeback": 18,
        "bandeira": 16,
        "valor_taxa": 14,
        "transportadora": 22,
        "numero_rastreio": 24,
        "data_envio_cliente": 26,
        "retorno_aprovacao": 22,
        "id_interno": 16,
        "motivo_recusa": 28,
        "obs": 36,
        "acao": 26,
        "aprovacao_por_quem": 24,
        "classificacao_ia": 24,
        "risco": 14,
        "pendencias": 36,
        "proxima_acao": 30,
        "status_prazo": 22,
        "status_vtex": 24,
        "data_faturamento": 18,
        "data_analise": 16,
        "origem_classificacao": 26,
    }
    for field, col_index in mapping.items():
        ws.column_dimensions[get_column_letter(col_index)].width = widths.get(field, 18)

    date_fields = [
        "data_transacao",
        "data_abertura_chargeback",
        "prazo_contestacao",
        "data_envio_cliente",
        "data_faturamento",
        "data_analise",
    ]
    for field in date_fields:
        if field in mapping:
            letter = get_column_letter(mapping[field])
            for cell in ws[f"{letter}{header_row + 1}:{letter}{ws.max_row}"]:
                cell[0].number_format = "dd/mm/yyyy"
    for field in ("valor_chargeback", "valor_taxa"):
        if field in mapping:
            letter = get_column_letter(mapping[field])
            for cell in ws[f"{letter}{header_row + 1}:{letter}{ws.max_row}"]:
                cell[0].number_format = 'R$ #,##0.00'


def export_csv(path: Path, ws, header_row: int, mapping: dict[str, int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ordered_fields = [field_for_header(header) for header in ORIGINAL_COLUMNS + GENERATED_COLUMNS]
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file, delimiter=";")
        writer.writerow(ORIGINAL_COLUMNS + GENERATED_COLUMNS)
        for row_index in range(header_row + 1, ws.max_row + 1):
            if empty_data_row(ws, row_index, mapping):
                continue
            output_row = []
            for field in ordered_fields:
                col_index = mapping.get(field) if field else None
                value = ws.cell(row=row_index, column=col_index).value if col_index else None
                if isinstance(value, (date, datetime)):
                    value = value.strftime("%d/%m/%Y")
                output_row.append(value)
            writer.writerow(output_row)


def make_backup(path: Path, backup_dir: Path) -> Path | None:
    if not path.exists():
        return None
    backup_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"{path.stem}_backup_{stamp}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def process_workbook(
    workbook_path: Path,
    import_records: list[dict[str, Any]] | None,
    output_dir: Path,
    sheet_name: str | None,
    export_format: str,
    backup: bool,
    near_days: int,
    today: date,
) -> dict[str, Any]:
    try:
        workbook = open_workbook(workbook_path)
    except PermissionError as exc:
        raise workbook_permission_error(workbook_path) from exc

    ws = choose_data_sheet(workbook, requested=sheet_name)
    header_row = find_header_row(ws)
    ensure_columns(ws, header_row)
    if import_records:
        append_records(ws, header_row, import_records)

    rows = classify_sheet(ws, header_row=header_row, near_days=near_days, today=today)
    mapping = ensure_columns(ws, header_row)
    apply_layout(ws, mapping=mapping, header_row=header_row, workbook=workbook)
    build_summary_sheet(workbook, rows)
    build_checklist_sheet(workbook, rows, today=today)
    build_lists_sheet(workbook)

    backup_path = make_backup(workbook_path, output_dir / "backups") if backup else None
    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(workbook_path)
    except PermissionError as exc:
        raise workbook_permission_error(workbook_path) from exc

    today_text = today.strftime("%Y-%m-%d")
    exports: list[Path] = []
    if export_format in {"xlsx", "both"}:
        xlsx_path = output_dir / f"chargeback_checklist_export_{today_text}.xlsx"
        workbook.save(xlsx_path)
        exports.append(xlsx_path)
    if export_format in {"csv", "both"}:
        csv_path = output_dir / f"chargeback_checklist_export_{today_text}.csv"
        export_csv(csv_path, ws, header_row=header_row, mapping=mapping)
        exports.append(csv_path)

    workbook.close()
    return {
        "rows": rows,
        "total": len(rows),
        "backup": backup_path,
        "exports": exports,
        "sheet": ws.title,
        "workbook": workbook_path,
    }
